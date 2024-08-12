import matplotlib.pyplot as plt
import numpy as np
import copy
import serial
import time
import math
import openpyxl
import os

# parameters
# length/e=0.001, e=mm
err=1e-4
d_platform=0.1
layer_height=0.28
ratio_e_xy=0.046564
drawback=6.5
z0=110
z_init=106

nx=4
ny=4
lx=51
ly=51
gap=lx-50


# file in/output

def get_gcode(filename):
    """
    Turn the gco. file into a list
    """
    with open(filename, 'r') as f_gcode:
        data = f_gcode.read().split("\n")
    return data

def output_gcode(filename,gcode_list):
    '''
    save a list as a gco. file
    '''
    filepath='/'.join(filename.split('/')[:-1])
    if os.path.isdir(filepath) is False:
        os.makedirs(filepath)
    if not gcode_list:
        print("output_gcode: No gcode")
        return
    with open(filename, 'w') as f_gcode:
        for ele in gcode_list:
            f_gcode.write(ele+"\n")
    return

# edit the excel with new experiment data
def output_excel(filename,parameters,new_value=None):
    book = openpyxl.load_workbook(filename)
    sheet = book.active
    last_row=len(sheet['A'])
    while sheet.cell(row=last_row,column=1).value is None:
        last_row=last_row-1
    sheet.cell(row=last_row+1,column=1).value=int(last_row-1)
    for i,ele in enumerate(parameters):
        sheet.cell(row=last_row+1,column=i+2).value=float(ele)
    if new_value:
        for jele in new_value.split(','):
            sheet['{0}{1}'.format(jele[0],last_row+1)].value=float(jele[1:])
    else:
        sheet['F{0}'.format(last_row+1)].value='-'

    book.save(filename)

# function used in 3d print and arduino

# 
def serWrite(ser,order):
    ser.write((order+'\r\n').encode('ascii'))
    time.sleep(1)

# function used in trim process

def nozzle_position_update(noz_end_pos=[0,0,0,0],gcode=None):
    # update nozzle position
    noz_init_pos=copy.deepcopy(noz_end_pos)
    g_point=gcode.split(' ')
    for igp in range(len(g_point)):
        if ';' in g_point[igp]:
            break
        if 'X' in g_point[igp]:
            noz_end_pos[0]=float(g_point[igp][1:])
        if 'Y' in g_point[igp]:
            noz_end_pos[1]=float(g_point[igp][1:])
        if 'Z' in g_point[igp]:
            noz_end_pos[2]=float(g_point[igp][1:])
        if 'E' in g_point[igp]:
            noz_end_pos[3]=float(g_point[igp][1:])
    return noz_init_pos,noz_end_pos

def line_cross2(p1,p2,lx=lx,ly=ly):
    points=[]

    # define the slope of the line from p1 to p2
    m=float('inf') if abs(p1[0]-p2[0])<err else (p1[1]-p2[1])/(p1[0]-p2[0])
    def funcy(ix):
        return (ix-p1[0])*m+p1[1]
    def funcx(iy):
        return (iy-p1[1])/m+p1[0]
    
    # let the points sorted by the value of x
    (px1,px2)=(p2,p1) if p1[0]>p2[0] else (p1,p2)
    
    # generate the intersections along the x-axis
    xcoor=[]
    for ix in range(math.floor((px1[0]-gap)/lx)-1,math.ceil((px2[0]+gap)/lx)+1):
        xi=ix*lx
        xcoor.extend([xi-gap,xi+gap])
    points.extend([[ele_x,funcy(ele_x)]for ele_x in xcoor if (px1[0]-ele_x)*(px2[0]-ele_x)<0])

    # let the points sorted by the value of y
    (py1,py2)=(p2,p1) if p1[1]>p2[1] else (p1,p2)

    # generate the intersections along the y-axis
    ycoor=[]
    for iy in range(math.floor((py1[1]-gap)/ly),math.ceil((py2[1]+gap)/ly)):
        yi=iy*ly
        ycoor.extend([yi-gap,yi+gap])
    points.extend([[funcx(ele_y),ele_y]for ele_y in ycoor if (py1[1]-ele_y)*(py2[1]-ele_y)<0])


    def same_platform(p,l,gap):
        return p[0]%l>gap and p[0]%l<l-gap and p[1]%l>gap and p[1]%l<l-gap
    if same_platform(p1,lx,gap):
        points.append(p1)
    if same_platform(p2,ly,gap):
        points.append(p2)

    sorted_points=sorted(points, key=lambda points: points[0]) if abs(p1[0]-p2[0])>err else sorted(points, key=lambda points: points[1])
    if p1[0]-p2[0]>err or p1[1]-p2[1]>err:
        sorted_points.reverse()

    return sorted_points

def gcode_raise(coor_orig,z_target,coor_comp=[5,49,56.3],suck_comp=-2):
    # 25,25,111.3 32,77,137.8
    # coor_comp = distance between magnetic and nozzle
    # coor_comp[2]-suck-comp[3]=comp for suck and for lift

    plane_height_comp=[
        [0.4,0.7,0.4,0.4],
        [-0.4,-0.2,0.2,-0.4],
        [-0.2,0.0,-0.2,0.2],
        [-0.2,-0.4,-0.2,0.4],
        ]
    plane_side_comp=[
        [10,0,0,0],
        [10,10,10,10],
        [10,10,10,10],
        [10,10,10,0]
        ]

    raise_gcode=[
        # move to the point above the certain platform unit
        "G0 F9000 X{0} Y{1} Z{2}".format(coor_orig[0]+coor_comp[0]+plane_side_comp[math.floor(coor_orig[0]/lx)][math.floor(coor_orig[1]/ly)],coor_orig[1]+coor_comp[1],z_target+coor_comp[2]+20),
        # send message "down" back to pc
        "M0 P500",
        "M118 A1 down",
        # 1.5s wait for motor movement
        "M0 S1.5",
        "G0 F9000 Z{0}".format(coor_orig[2]+coor_comp[2]+suck_comp),
        "G0 F9000 Z{0}".format(z_target+coor_comp[2]+plane_height_comp[math.floor(coor_orig[0]/lx)][math.floor(coor_orig[1]/ly)]),
        "M0 P500",
        "M118 A1 up",
        "M0 S1.5"
        ]
    return raise_gcode

def fix_gcode(gcode_line,fix_para_arr):
    gcode_unit_arr=gcode_line.split(' ')  
    for i,ele in enumerate(gcode_unit_arr):
        startword=ele[0] if ele[0] else None
        match startword:
            case ';':
                break
            case 'X'|'Y'|'Z'|'E'|'G'|'F':
                fix_number=[e[1:] for e in fix_para_arr if startword in e]
                if fix_number:
                    if abs(float(fix_number[0])+1)<err:
                        gcode_unit_arr[i]=''
                        continue
                    gcode_unit_arr[i]='{0}{1:.5f}'.format(startword,float(fix_number[0])) if startword != 'G' else '{0}{1}'.format(startword,int(fix_number[0]))
            case _:
                pass
                
    fixed_gcode=" ".join(gcode_unit_arr)
    return fixed_gcode

# function varify

def gcode_visualize(filename,targert_layer_num):
    gcode=get_gcode(filename)

    fig = plt.figure()
    
    ax = fig.add_subplot(111, projection='3d')

    ax.set_xlim3d([0, 300])
    ax.set_ylim3d([0, 300])
    ax.set_zlim3d([0, 300])

    noz_init_pos=[0,0,0]
    noz_end_pos=[0,0,0]

    start_print=False
    for i,g_line in enumerate(gcode):
        if 'LAYER:{}'.format(int(targert_layer_num)) in g_line:
            start_print=True
        if 'G91' in g_line or ';MESH:Template_ladder.stl' in g_line:
            break
        if not start_print:
            continue
        if 'G0' in g_line or 'G1' in g_line:
            noz_init_pos=copy.deepcopy(noz_end_pos)
            g_point=g_line.split(' ')
            for ele in g_point:
                if ';' in ele:
                    break
                if 'X' in ele:
                    noz_end_pos[0]=float(ele[1:])
                if 'Y' in ele:
                    noz_end_pos[1]=float(ele[1:])
                if 'Z' in ele:
                    noz_end_pos[2]=float(ele[1:])
            if noz_init_pos[0]==0:
                continue
            if 'G1' in g_line:
                ax.plot([noz_init_pos[0],noz_end_pos[0]],[noz_init_pos[1],noz_end_pos[1]],[noz_init_pos[2],noz_end_pos[2]])
                
                ax.text(noz_init_pos[0],noz_init_pos[1],noz_init_pos[2],'{}'.format(g_line))
        print(g_line)
        plt.pause(0.1)
    plt.show()

    # module.output_gcode("test.gcode",gcode)


# whole process

def platform_reset(platform_matrix,z_init=z_init,z_target=z0):

    gcode_reset=['G28',
                 'G92 E0',
                 'G0 Z120',
                 "G0 F9000 X{0:.5f} Y{1:.5f} Z{2:.5f}".format(int(nx/2)*lx,int(ny/2)*ly,z0+10),
                 'M0']
    
    for i in range(nx):
        for j in range(ny-1,-1,-1):
            if platform_matrix[i][j]==0:
                continue
            gcode_reset.extend(gcode_raise([(i+0.5)*lx,(j+0.5)*ly,z_init],z_target))
    return gcode_reset

def prefix_Translation_HeightDetection(gcode,x_comp=20,y_comp=20,z_comp=z0,theta=math.pi/4):
    
    # initialize parameters of height detect
    platform_matrix=[[0]*nx for i in range(ny)]
    noz_init_pos=[0,0,0,0]
    noz_end_pos=[0,0,0,0]

    # initialize the center to rotate by
    model_edge_pos=[]

    start_comp=False
    type_sup=False
    for i_line,ele_gcode in enumerate(gcode):

        if 'G91' in ele_gcode:
            break

        if ';MINX' in ele_gcode or ';MINY' in ele_gcode or ';MAXX' in ele_gcode or ';MAXY' in ele_gcode:
            model_edge_pos.append(float(ele_gcode.split(':')[-1]))

        if ';LAYER:0' in ele_gcode:
            start_comp=True
            cx,cy=(model_edge_pos[0]+model_edge_pos[2])/2,(model_edge_pos[1]+model_edge_pos[3])/2
            rotate_matrix=np.array([[math.cos(theta) ,-math.sin(theta)],
                                    [math.sin(theta),math.cos(theta)]])
            pos_matrix=np.array([0,1])

        if 'TYPE' in ele_gcode:
            type_sup=True if ('SUP' in ele_gcode or 'SKIRT' in ele_gcode) and 'INTERFACE' not in ele_gcode else False

        if start_comp and ele_gcode.startswith('G') and ('X' in ele_gcode or 'Y' in ele_gcode):
            noz_init_pos,noz_end_pos=nozzle_position_update(noz_end_pos,ele_gcode)

            # position matrix is the relative position to rotation center
            pos_matrix=np.array([np.float64(noz_end_pos[0])-cx,np.float64(noz_end_pos[1])-cy])
            new_pos_matrix=np.matmul(rotate_matrix,pos_matrix)

            noz_end_pos[0]=new_pos_matrix[0]+cx+x_comp
            noz_end_pos[1]=new_pos_matrix[1]+cy+y_comp

            # edge detection
            if noz_end_pos[0]>nx*lx or noz_end_pos[1]>ny*ly or noz_end_pos[0]<0 or noz_end_pos[1]<0:
                return None,None

            # update the original gcode
            gcode[i_line]=fix_gcode(ele_gcode,['X{:.5f}'.format(noz_end_pos[0]),
                                            'Y{:.5f}'.format(noz_end_pos[1]),
                                            'Z{:.5f}'.format(noz_end_pos[2]+z_comp)])
                
            if not type_sup and 'G1' in ele_gcode and 'X' in ele_gcode:
                p_list=line_cross2(noz_init_pos[:2],noz_end_pos[:2])
                for ele_plist in p_list:
                    for dx in [-gap,0,gap]:
                        for dy in [-gap,0,gap]:
                            if (ele_plist[0]+dx)/lx>=nx or (ele_plist[1]+dy)/ly>=ny or (ele_plist[0]+dx)/lx<0 or (ele_plist[1]+dy)/ly<0:
                                continue
                            if platform_matrix[int((ele_plist[0]+dx)/lx)][int((ele_plist[1]+dy)/ly)]==0:
                                platform_matrix[int((ele_plist[0]+dx)/lx)][int((ele_plist[1]+dy)/ly)]=noz_end_pos[2]+z_comp                   

    return gcode,platform_matrix

def suffix_Mark(gcode,parameters=[-1,-1,-1,-1]):
    type_fill=False
    new_gcode=['M104 S200','G28']

    for i_gcode,ele_gcode in enumerate(gcode):
        new_gcode.append(ele_gcode)
        if ';platform raise' in ele_gcode:
            new_gcode.extend(['M0 P100','M118 A1 pause1','M0','M0 P100','M118 A1 pause2'])
        if 'LAYER:0' in ele_gcode:
            new_gcode.extend(['M0 P100','M118 A1 parameters_{0}_{1:.5f}_{2:.5f}_{3:.5f}'.format(parameters[0],parameters[1],parameters[2],parameters[3])])
            new_gcode.extend(['M0 P100','M118 A1 time_start'])
        if 'G91' in ele_gcode:
            new_gcode.extend(['M0 P100','M118 A1 time_end'])
        if 'FILL' in ele_gcode or 'WALL' in ele_gcode or 'SKIN' in ele_gcode:
            if not type_fill:
                new_gcode.extend(['M0 P100','M118 A1 time_fill_start'])
            type_fill=True
        if 'NONMESH' in ele_gcode or 'TIME_ELAPSED' in ele_gcode:
            if type_fill:
                new_gcode.extend(['M0 P100','M118 A1 time_fill_end'])
            type_fill=False
    return new_gcode

def prefix_Trim3(gcode,platform_matrix):

    new_gcode=[]
    # template parameter used in loop
    start_print=False
    type_sup=False
    noz_init_pos=[0,0,0,0]
    noz_end_pos=[0,0,0,0]

    trimmed_e=0
    total_trim_e=0
    for i,g_line in enumerate(gcode):

        # checkpoint at certain layer
        if ';LAYER:1' in g_line:
            pass

        # start after layer:0
        if ';LAYER:0' in g_line:
            start_print = True

        # end the print
        if 'G91' in g_line:
            start_print = False
            total_trim_e=total_trim_e+trimmed_e

        # choose sup and skirt to trim
        # type sup > mesh > type wall
        if 'TYPE' in g_line or 'MESH' in g_line:       
            if ('SUP' in g_line or 'SKIRT' in g_line) and 'INTERFACE' not in g_line :
                type_sup=True
            else:
                type_sup=False

        # gcode for lift the layer during the print
        if ';LAYER:' in g_line and noz_init_pos[2] !=0:
            for xi in range(len(platform_matrix)):
                for yi in range(len(platform_matrix[xi])):
                    if abs(platform_matrix[xi][yi]-noz_init_pos[2])<layer_height/2:
                        # new_gcode.append('M0')
                        new_gcode.extend([';platform raise'])
                        new_gcode.extend(gcode_raise([(xi+0.5)*lx,(yi+0.5)*ly,z0],noz_init_pos[2]-layer_height))
            new_gcode.append('G0 F9000 X{0} Y{1} Z{2}'.format(noz_init_pos[0],noz_init_pos[1],noz_init_pos[2]))
        
        # main part to trim sup
        if start_print and g_line.find('G')==0:

            if 'G92 E0' in g_line:
                total_trim_e=total_trim_e+trimmed_e
                trimmed_e=0
            # update nozzle position
            noz_init_pos,noz_end_pos=nozzle_position_update(noz_end_pos,g_line)
            # detect which part to trim
            # new point would be p_ele
            if type_sup:
                # deal with orig drawback in type sup
                if 'E' in g_line and 'X' not in g_line:
                    new_gcode.append(fix_gcode(g_line,['E{0:.5f}'.format(noz_end_pos[3]-trimmed_e)])+";orig drawback")
                    continue
                # enter two point and get point seperate by platform
                p_list=line_cross2(noz_init_pos[:2],noz_end_pos[:2])

                e_real=0
                e_feed=0
                # skip the part 
                # trim one gcode line from origs
                for ip in range(0,len(p_list)-1,2):
                    # skip the points below the raised platform
                    if platform_matrix[int(p_list[ip][0]/lx)][int(p_list[ip][1]/ly)]>noz_end_pos[2]:
                        continue

                    dist_xy=math.dist(p_list[ip],p_list[ip+1])
                    # e_real is sum of the real extrusion, without the extrusion removed
                    e_real=e_real+dist_xy*ratio_e_xy
                    
                    # skip continuous G0
                    if 'G1' in g_line:
                        e_feed=noz_init_pos[3]-trimmed_e+e_real
                        
                        new_gcode.append("G0 F9000 X{0:.5f} Y{1:.5f}".format(p_list[ip][0],p_list[ip][1]))
                        new_gcode.append("G1 F1500 X{0:.5f} Y{1:.5f} E{2:.5f}".format(p_list[ip+1][0],p_list[ip+1][1],e_feed))
                
                # append new points back to gcode
                # trimmed e = orig E diff - real E diff
                d_trimmed_e=noz_end_pos[3]-noz_init_pos[3]-e_real
                # if d_trimmmed_e < 1e-4 it belongs to the calculate error, ignore it
                trimmed_e=trimmed_e+(d_trimmed_e if d_trimmed_e > err else 0)
            # not sup part,include fill, wall, etc...
            # only need to minus a constant E
            elif 'E' in g_line:
                new_gcode.append(fix_gcode(g_line,['E{:.5f}'.format(noz_end_pos[3]-trimmed_e)]))
            else:
                new_gcode.append(g_line)      
        else:
            new_gcode.append(g_line)
    
    return new_gcode,total_trim_e

def suffix_Separate3(gcode):

    new_gcode=[]
    type_sup=False
    be_drawback_platform=True
    noz_end_pos=[0,0,0,0]
    ele_noz_init_pos=[0,0,0,0]
    ele_noz_end_pos=[0,0,0,0]

    e0_sup_same_layer=0

    temp_sup_gcode=[[[],[],[],[]],
                    [[],[],[],[]],
                    [[],[],[],[]],
                    [[],[],[],[]]]

    for i_gcode_line,gcode_line in enumerate(gcode):

        if 'G' in gcode_line:
            _,noz_end_pos=nozzle_position_update(noz_end_pos,gcode_line)

        if 'TYPE' in gcode_line:
            if ('SUP' in gcode_line or 'SKIRT' in gcode_line) and 'INTERFACE' not in gcode_line :
                if not type_sup:           
                    ele_noz_end_pos=copy.deepcopy(noz_end_pos)
                    ele_noz_end_pos[3]=ele_noz_end_pos[3]+6.5
                    e0_sup_same_layer=ele_noz_end_pos[3]
                    pass
                    # e_comp_sum=noz_end_pos[3]+6.5
                type_sup=True
                # e_comp=noz_end_pos[3]+6.5

            else:
                # sum drawback during support of same layer
                # append drawback between platform
                if type_sup:
                    # new_gcode.append('G1 F1500 E{0:.5f};e0_sup_same_layer_return'.format(e0_sup_same_layer))
                    for i_p in range(nx):
                        for j_p in range(nx):
                            if any(temp_sup_gcode[i_p][j_p]):
                                if 'G1' in temp_sup_gcode[i_p][j_p][0]:
                                    temp_sup_gcode[i_p][j_p][0]=fix_gcode(temp_sup_gcode[i_p][j_p][0],['G0','F9000','E-1'])
                                for ele_temp_sup in temp_sup_gcode[i_p][j_p]:
                                    if 'E' in ele_temp_sup:
                                        if be_drawback_platform:
                                            new_gcode.append('G1 F1500 E{0:.5f};switch_platform_return'.format(e0_sup_same_layer))
                                            be_drawback_platform=False
                                        _,temp_noz_poz=nozzle_position_update(gcode=ele_temp_sup)
                                        e0_sup_same_layer=e0_sup_same_layer+temp_noz_poz[3]
                                        new_gcode.append(fix_gcode(ele_temp_sup,['E{0:.5f}'.format(e0_sup_same_layer)]))
                                    else:
                                        new_gcode.append(ele_temp_sup)
                                if not be_drawback_platform:
                                    new_gcode.append('G1 F1500 E{0:.5f};switch_platform_drawback'.format(e0_sup_same_layer-drawback))
                                    be_drawback_platform=True
                            # new_gcode.extend(temp_sup_gcode[i_p][j_p])
                            
                    # new_gcode.append('G1 F1500 E{0:.5f};e0_sup_same_layer_drawback'.format(e0_sup_same_layer-drawback))
                    
                    # return to the position of next not sup type
                    new_gcode.append('G0 F9000 X{0:.5f} Y{1:.5f};last_sup_point'.format(ele_noz_end_pos[0],ele_noz_end_pos[1]))
                    

                    # reset temp_sup_gcode after adding the support of a layer
                    temp_sup_gcode=[[[],[],[],[]],
                                    [[],[],[],[]],
                                    [[],[],[],[]],
                                    [[],[],[],[]]]
                type_sup=False


        if type_sup and 'G' in gcode_line:

            # skip the G0 series (skip move only)
            if ('G0' in gcode_line and 'Z' not in gcode_line)and('G0' in gcode[i_gcode_line+1] and 'Z' not in gcode[i_gcode_line+1]):
                continue

            # skip all drawback
            if ('G1' in gcode_line  and 'X' not in gcode_line):
                continue

            ele_noz_init_pos,ele_noz_end_pos=nozzle_position_update(ele_noz_end_pos,gcode_line)

            gcode_line=fix_gcode(gcode_line,["E{0:.5f}".format(ele_noz_end_pos[3]-ele_noz_init_pos[3])])

            temp_sup_gcode[int(ele_noz_end_pos[0]/lx)][int(ele_noz_end_pos[1]/ly)].append(gcode_line)
            
            continue

        new_gcode.append(gcode_line)
    # print(temp_sup_gcode)  

    return new_gcode
    # output_gcode('test_suffix_separate_new_.gcode',new_gcode)



'''

def preprocess_simply(filename,x_comp=20,y_comp=20,z_comp=z0,theta=math.pi/4,nx=nx,ny=ny,lx=lx,ly=ly):

    gcode=get_gcode(filename)

    # initialize parameters of height detect
    platform_matrix=[[0]*nx for i in range(ny)]

    noz_init_pos=[0,0,0,0]
    noz_end_pos=[0,0,0,0]

    # initialize the center to rotate by
    model_edge_pos=[]

    start_comp=False
    type_sup=False
    for i_line,g_line in enumerate(gcode):

        if 'G91' in g_line:
            break

        if ';MINX' in g_line or ';MINY' in g_line or ';MAXX' in g_line or ';MAXY' in g_line:
            model_edge_pos.append(float(g_line.split(':')[-1]))

        if ';LAYER:0' in g_line:
            start_comp=True
            cx,cy=(model_edge_pos[0]+model_edge_pos[2])/2,(model_edge_pos[1]+model_edge_pos[3])/2
            rotate_matrix=np.array([[math.cos(theta) ,-math.sin(theta)],
                                    [math.sin(theta),math.cos(theta)]])
            pos_matrix=np.array([0,1])

        if 'TYPE' in g_line:
            type_sup=type_sup_or_not(g_line)

        if start_comp and g_line.find('G')==0 and ('X' in g_line or 'Y' in g_line):
            noz_init_pos,noz_end_pos=nozzle_position_update(noz_end_pos,g_line)

            # position matrix is the relative position to rotation center
            pos_matrix=np.array([np.float64(noz_end_pos[0])-cx,np.float64(noz_end_pos[1])-cy])
            new_pos_matrix=np.matmul(rotate_matrix,pos_matrix)
            
            # update the original gcode
            g_point=g_line.split(' ')  
            for i_g_p in range(len(g_point)):
                if g_point[i_g_p].find(';')==0:
                    break
                if 'X' in g_point[i_g_p]:
                    noz_end_pos[0]=new_pos_matrix[0]+cx+x_comp
                    g_point[i_g_p]='X{:.5f}'.format(noz_end_pos[0])
                if 'Y' in g_point[i_g_p]:
                    noz_end_pos[1]=new_pos_matrix[1]+cy+y_comp
                    g_point[i_g_p]='Y{:.5f}'.format(noz_end_pos[1])
                if g_point[i_g_p].find('Z')==0:
                    noz_end_pos[2]=noz_end_pos[2]+z0
                    g_point[i_g_p]="Z{:.5f}".format(noz_end_pos[2])
            
            gcode[i_line]=" ".join(g_point)

            if not type_sup and 'G1' in g_line and 'X' in g_line:
                p_list=line_cross(noz_init_pos[:2],noz_end_pos[:2])
                for ip in range(len(p_list)-1):
                    point_center=midpoint(p_list[ip],p_list[ip+1])
                    if closest_diff(point_center[0],lx)<err:
                        continue
                    if platform_matrix[int(point_center[0]/lx)][int(point_center[1]/ly)]==0:
                        platform_matrix[int(point_center[0]/lx)][int(point_center[1]/ly)]=noz_end_pos[2]

    # newfilename=filename.split('.')[0]+'_new.gcode'

    # output_gcode(newfilename,gcode)

    return platform_matrix,gcode



'''